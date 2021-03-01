# -*- coding: utf-8 -*-
# author: fanchuang
# DateTime:2021/2/23 0023 17:33 
# contact: fanchuangwater@gmail.com
# about: 把数据库的文件写入到 Excel


from datetime import datetime
from openpyxl import Workbook
from utils.mongoDB import Mongo
from config import books_db, book_urls_coll


def write_excel():
    excel_name = 'Amazon_Computer_Book_Urls.xlsx'
    sheet_name = 'book_urls'
    client = Mongo(books_db, book_urls_coll)

    wb = Workbook()
    if sheet_name not in wb.sheetnames:
        working_sheet = wb.create_sheet(sheet_name, 0)  # 添加到开头
    else:
        working_sheet = wb[sheet_name]

    working_sheet["A1"] = "序号"
    working_sheet["B1"] = "ID"
    working_sheet["C1"] = "链接"
    working_sheet["D1"] = "生成日期"

    index = 1
    for x in client.coll.find():
        working_sheet[f'A{index + 1}'] = index
        working_sheet[f'B{index + 1}'] = str(x['_id'])
        working_sheet[f'C{index + 1}'] = x['url']
        working_sheet[f'D{index + 1}'] = str(datetime.now()).split(".")[0]
        index += 1

    wb.save(excel_name)


def write_all_books():
    excel_name = 'Amazon_Computer_Books_Info.xlsx'
    sheet_name = 'book_urls'
    # client = Mongo(books_db, book_urls_coll)
    # client = Mongo("amazon_books", "books_info_9")    #
    client = Mongo("amazon_books", "books_info_10")

    wb = Workbook()
    if sheet_name not in wb.sheetnames:
        working_sheet = wb.create_sheet(sheet_name, 0)  # 添加到开头
    else:
        working_sheet = wb[sheet_name]

    working_sheet["A1"] = "序号"
    working_sheet["B1"] = "书名"
    working_sheet["C1"] = "作者"
    working_sheet["D1"] = "价格"

    working_sheet["E1"] = "内容简介"
    working_sheet["F1"] = "推荐理由"
    working_sheet["G1"] = "作者简介"

    working_sheet["H1"] = "ASIN"
    working_sheet["I1"] = "出版社"
    working_sheet["J1"] = "出版日期"
    working_sheet["K1"] = "品牌"
    working_sheet["L1"] = "语言"
    working_sheet["M1"] = "文件大小"

    working_sheet["N1"] = "标准语音朗读"
    working_sheet["O1"] = "X-Ray"
    working_sheet["P1"] = "纸书页数"
    working_sheet["Q1"] = "封面链接"
    working_sheet["R1"] = "图书链接"
    working_sheet["S1"] = "生词提示功能"


    index = 1
    for x in client.coll.find():
        working_sheet[f'A{index + 1}'] = index
        working_sheet[f"B{index + 1}"] = x["book_name"]
        working_sheet[f"C{index + 1}"] = x["author"]
        working_sheet[f"D{index + 1}"] = x["price"]

        working_sheet[f"E{index + 1}"] = x["book_about"]
        working_sheet[f"F{index + 1}"] = x["recommend_reason"]
        working_sheet[f"G{index + 1}"] = x["author_info"]

        working_sheet[f"H{index + 1}"] = x["ASIN"]
        working_sheet[f"I{index + 1}"] = x["publisher"]
        working_sheet[f"J{index + 1}"] = x["publish_time"]
        working_sheet[f"K{index + 1}"] = x["brand"]
        working_sheet[f"L{index + 1}"] = x["language"]
        working_sheet[f"M{index + 1}"] = x["file_size"]
        working_sheet[f"N{index + 1}"] = x["voice_status"]

        working_sheet[f"O{index + 1}"] = x["X-Ray"]
        working_sheet[f"P{index + 1}"] = x["pages"]
        working_sheet[f"Q{index + 1}"] = x["book_image_url"]
        working_sheet[f"R{index + 1}"] = x["book_url"]
        working_sheet[f"S{index + 1}"] = x["reading_helper"]

        index += 1

    wb.save(excel_name)


if __name__ == '__main__':
    # write_excel()
    write_all_books()

